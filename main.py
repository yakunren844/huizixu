import math
import os
from datetime import datetime
from time import sleep
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from playwright.sync_api import sync_playwright
import shutil

# 你的登录用户名
username="wangu_02"
# 你的登录密码
password="wangu@123456"

# 需要填写的每天限制信号量
xinhao='30'
#需要填写的到期日期，格式是yyyy-mm-dd,例如：1990-01-01
limit_day="2026-08-19"
#需要改成你需要处理文件名，如果没有放在本程序目录下，需要完整路径
deal_file_name="需要分割.xlsx"         # 需要分割的excel文件
deal_file_limit=5000            # 需要分割的excel文件的行数,默认为5000
deal_file_path=r"D:\\develop\\huizixu\\20250821提报"        # 需要上传的excel文件夹
deal_type=2                     # 1是自动分割大excel文件，然后自动上传
                                # 2是处理已经分割好的文件，直接上传 

######################################################################
                #           _ooOoo_
                #     　　 o8888888o
                #     　　 88" . "88
                #     　　 (| -_- |)
                #     　　  O\ = /O
                #     　　 _/`---'\_
                #     　 .' \\| |// '.
                #     　 / \\||| : |||// \
                #      / _||||| -:- |||||- \
                #     | | \\\ - / | |
                #     | \_| ''\---/'' |_/ |
                #     \ .-\__ `-` ___/-. /
                #    ___`..' /--.--\ `.. __
                # . "" '< `.___\_<|>_/___.' >' "".
                # | | : `- \`.;`\ _ /`;.`/ - ` : | |
                # \ \ `-. \_ __\ /__ _/ .-` / /
######################################################################
#下面不用动
browser=None


def upload_file_to_website(page, base_path,file_name):
    cur_path=os.path.join(base_path, "excel", file_name)
    file_success_path=os.path.join(base_path, "success", file_name)
    file_failed_path=os.path.join(base_path, "failed", file_name)
    
    page.goto('https://www.likewon.com/userNew/monitor/addList', timeout=30000)
    page.set_input_files('input[type="file"]', cur_path, timeout=600000)
    # page.click('//*[@id="pane-addList"]/div/div[4]/form/div[1]/div/div/div/button[1]/span')
    page.click('//*[@id="pane-addList"]/div/div[4]/form/div[2]/div/div/div/div/input', timeout=600000) #显示信号量
    # page.click('//html/body/div[3]/div[1]/div[1]/ul/li[1]/span')  #选择信号量
    # 2. 选择选项（等待下拉菜单出现）
    page.locator(".el-select-dropdown__item:has-text('m10/普通精准')").click()

    page.fill('//*[@id="pane-addList"]/div/div[4]/form/div[3]/div/div[1]/input', xinhao)
    page.fill('//*[@id="pane-addList"]/div/div[4]/form/div[4]/div/div/input', limit_day)
    page.click('//*[@id="pane-addList"]/div/div[4]/form/button')
    # page.click('/html/body/div[5]/div/p')
    sleep(3)
    
    try:
        page.wait_for_selector("p:has-text('加载中')", state="detached", timeout=60000)
        sleep(1)
        # page.wait_for_selector("p:has-text('上传成功！')", timeout=3000, state="visible")
        print(f"{file_name}文件上传成功！")
        shutil.move(cur_path, file_success_path)
        return True
    except:
        try:
            page.wait_for_selector('//*[@id="pane-addList"]/div/div[5]/div/div[3]/span/button', timeout=1000, state="visible")
            page.click('//*[@id="pane-addList"]/div/div[5]/div/div[1]/button')
            # page.wait_for_selector('//*[@id="pane-addList"]/div/div[5]/div/div[3]/span/button', timeout=1000, state="visible")
            # page.click('//*[@id="pane-addList"]/div/div[5]/div/div[3]/span/button')
            print(f"{file_name}文件上传成功,有重复名单")
            shutil.move(cur_path, file_success_path)
            return True
        except:
            shutil.move(cur_path, file_failed_path)
            print(f"{file_name}文件上传失败！")
            return False
        

def login_to_website(  ):
    p=sync_playwright().start()
    # with sync_playwright() as p:
        # 启动浏览器（默认chromium，可改为 firefox 或 webkit）
    browser = p.chromium.launch(headless=False,slow_mo=1000)  # headless=True 无界面模式
    # 设置所有操作的默认延迟（毫秒）
    # 创建上下文时不带 slow_mo
    page = browser.new_page()  
    # page = browser.new_page()
    
    try:
        url='https://www.likewon.com/'
        # 导航到目标页面
        page.goto(url, timeout=30000)  # 10秒超时
        print(f"已访问: {url}")
        if page.query_selector('//*[@id="header"]/div[1]/div[2]/div[1]/ul/li[3]/a'):
            print("已经登陆完成")
            return None
        page.click('//*[@id="header"]/div[1]/div[2]/div/ul/li[2]/button/span')  # XPath
        page.click('id=tab-second')
        page.fill('input[placeholder="请输入登录名"]', username)
        page.fill('input[type="password"]', password)
        page.click('//*[@id="pane-second"]/div/form/div[3]/div/button/span')
        sleep(1)
        if page.query_selector('//*[@id="header"]/div[1]/div[2]/div[1]/ul/li[3]/a'):
            page.click('//*[@id="header"]/div[1]/div[2]/div[1]/ul/li[3]/a')
            print("已经登陆完成....")
            return page
            # upload_file_to_website(page, "D:\\develop\\huizixu\\temp_20250815_152953\\part_2.xlsx")

        # sleep(60000)

    except Exception as e:
        print(f"登录失败: {str(e)}")
        page.screenshot(path="login_failed.png")  # 失败时截图
        return None
    # finally:
    #     browser.close()
def deal_files(page,base_file_path):
    upload_file_number=0    #需要上传的文件数量
    failed_file_number=0    #上传失败的文件数量
    success_file_number=0    #上传成功的文件数量

    excel_file_path=os.path.join(base_file_path, "excel")
    for file in os.listdir(excel_file_path):
        if file.endswith(".xlsx"):
            upload_file_number+=1
            ret=upload_file_to_website(page, base_file_path, file) 
            if ret:
                success_file_number+=1
            else:
                failed_file_number+=1
    print(f"需要上传的文件数量：{upload_file_number}",f",上传成功：{success_file_number}",f",上传失败：{failed_file_number}")
   
def merge_and_rename_folders(source_dirs, target_dir, prefix="", digits=4):
    """
    将多个文件夹中的文件合并到一个目标文件夹，并按数字顺序重命名
    参数:
        source_dirs: 包含多个源文件夹路径的列表
        target_dir: 目标文件夹路径
        prefix: 重命名后的文件前缀(可选)
        digits: 文件名数字部分的位数(默认为4)
    """
    # 创建目标文件夹(如果不存在)
    os.makedirs(target_dir, exist_ok=True)
    
    # 计数器初始化
    count = 1
    
    # 遍历所有源文件夹
    for source_dir in source_dirs:
        # 确保源文件夹存在
        if not os.path.isdir(source_dir):
            print(f"警告: 源文件夹不存在 {source_dir}")
            continue
            
        # 遍历源文件夹中的所有文件
        for filename in os.listdir(source_dir):
            source_path = os.path.join(source_dir, filename)
            
            # 只处理文件(跳过子文件夹)
            if os.path.isfile(source_path):
                # 获取文件扩展名
                ext = os.path.splitext(filename)[1]
                
                # 生成新文件名(数字部分补零)
                new_filename = f"{prefix}{count:0{digits}d}{ext}"
                target_path = os.path.join(target_dir, new_filename)
                
                # 复制文件到目标位置
                shutil.copy2(source_path, target_path)
                print(f"已复制: {source_path} -> {target_path}")
                
                # 计数器增加
                count += 1
def split_excel_pd(input_file):
    """分割大Excel文件到临时文件夹"""
    upload_file_list=[]
    upload_file_path=""
     # 固定的标题头
    fixed_headers = ['监测名单', '证件号码', '类型']
    
    # 1. 创建带时间戳的临时文件夹
    upload_file_path = f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    os.makedirs(upload_file_path, exist_ok=True)
    os.makedirs(f"{upload_file_path}/excel", exist_ok=True)
    os.makedirs(f"{upload_file_path}/failed", exist_ok=True)
    os.makedirs(f"{upload_file_path}/success", exist_ok=True)
    print(f"创建文件夹: {upload_file_path}")
    try:
        # 读取 Excel 文件
        xls = pd.ExcelFile(input_file)

        # 遍历每个 sheet
        for sheet_name in xls.sheet_names:
            # 读取 sheet 数据
            df = pd.read_excel(xls, sheet_name=sheet_name)

            # 检查标题头
            if not all(header in df.columns for header in fixed_headers):
                print(f"Sheet '{sheet_name}' 缺少固定的标题头（监测名单、证件号码、类型），跳过")
                continue
            # 只保留固定的标题头列（忽略多余列）
            df = df[fixed_headers]
             # 如果 sheet 为空，仅保存标题头
            if df.empty:
                output_file = os.path.join(upload_file_path, 'excel', "part_0001.xlsx")
                df.to_excel(output_file, index=False, columns=fixed_headers)
                print(f"Sheet '{sheet_name}' 为空，生成文件: {output_file}")
                continue
             # 按 max_rows 分割数据
            total_rows = len(df)
            for start in range(0, total_rows, deal_file_limit):
                # 提取当前分片
                df_chunk = df.iloc[start:start + deal_file_limit]

                # 生成输出文件路径（part_0001, part_0002, ...）
                part_index = start // deal_file_limit + 1
                temp_name="part_"+str(part_index).zfill(3)+".xlsx"
                output_file = os.path.join(upload_file_path,"excel",temp_name)

                # 保存到新的 Excel 文件
                df_chunk.to_excel(output_file, index=False, columns=fixed_headers)
                upload_file_list.append(temp_name)
                print(f"生成文件: {output_file}，行数: {len(df_chunk)}")
        return upload_file_path,upload_file_list

    except Exception as e:
        
        print(f"发生错误: {e}")
        return None,None


def split_large_excel(input_file):
    """分割大Excel文件到临时文件夹"""
    upload_file_list=[]
    upload_file_path=""
    
    # 1. 创建带时间戳的临时文件夹
    upload_file_path = f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    os.makedirs(upload_file_path, exist_ok=True)
    os.makedirs(f"{upload_file_path}/excel", exist_ok=True)
    os.makedirs(f"{upload_file_path}/failed", exist_ok=True)
    os.makedirs(f"{upload_file_path}/success", exist_ok=True)
    print(f"创建文件夹: {upload_file_path}")

    # 2. 加载原始文件
    wb = load_workbook(input_file)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    # 2. 加载原始文件
    wb = load_workbook(input_file)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    
    # 3. 获取列头
    headers = [ws.cell(row=1, column=col).value for col in range(1, max_col+1)]
    
    # 4. 计算需要分割的文件数量
    total_chunks = math.ceil((max_row - 1) / deal_file_limit)    # -1是因为排除列头行
    print(f"总行数: {max_row} | 每份行数: {deal_file_limit} | 将分割为 {total_chunks} 个文件")

    # 5. 分割处理
    for chunk_num in range(total_chunks):
        # 创建新工作簿
        new_wb = load_workbook(input_file)
        new_ws = new_wb.active
        new_ws.delete_rows(2, new_ws.max_row)  # 清空模板中的数据，只保留列头

        # 计算行范围 (注意: 跳过原列头行)
        start_row = 2 + chunk_num * deal_file_limit
        end_row = min(start_row + deal_file_limit, max_row + 1)
        
        # 复制数据行
        for row in range(start_row, end_row):
            new_row = []
            for col in range(1, max_col+1):
                new_row.append(ws.cell(row=row, column=col).value)
            new_ws.append(new_row)
            
        file_number=chunk_num
        # 保存文件
        while True:
            file_number+=1
            temp_name="part_"+str(file_number).zfill(3)+".xlsx"
            output_file = os.path.join(upload_file_path, 'excel',temp_name)
            print(f"保存文件：{output_file}")
            if os.path.exists(output_file):
                print(f"文件已存在：{output_file}")
                continue
            else:
                print(f"保存文件：{output_file}")
                new_wb.save(output_file)
                upload_file_list.append(temp_name)
                print(f"已生成: {output_file} (行 {start_row-1}-{end_row-1})")
                break

    print(f"\n分割完成！所有文件已保存至: {os.path.abspath(upload_file_path)}")
    return upload_file_path,upload_file_list

def pre_upload_path(deal_file_path):
    '''
    处理文件夹
    '''
    if deal_file_path=="":
        print("处理文件夹，deal_file_path文件路径不能为空")
        exit()

    excel_file_path=os.path.join(deal_file_path, "excel")
    if not os.path.exists(excel_file_path):
        os.makedirs(excel_file_path)
        print("处理文件夹，excel文件夹不存在,创建excel文件夹")
    success_file_path=os.path.join(deal_file_path, "success")
    if not os.path.exists(success_file_path):
        os.makedirs(success_file_path)
        print("处理文件夹，success文件夹不存在，创建success文件夹")

    failed_file_path=os.path.join(deal_file_path, "failed")
    if not os.path.exists(failed_file_path):
        os.makedirs(failed_file_path)
        print("处理文件夹，failed文件夹不存在，创建failed文件夹")

    for file in os.listdir(deal_file_path):
        if file.endswith(".xlsx"):
            shutil.move(os.path.join(deal_file_path, file), os.path.join(excel_file_path, file))
    file_path=deal_file_path
    return file_path

# 使用示例
if __name__ == "__main__":
     # 示例用法
    # source_dirs = [
    #     "output\\output_1",
    #     "output\\output_2",
    #     "output\\output_3",
    #     "output\\output_4"
    # ]
    # target_dir = "output_ok"
    
    # # 调用函数(可以自定义前缀和位数)
    # merge_and_rename_folders(
    #     source_dirs=source_dirs,
    #     target_dir=target_dir,
    #     prefix="file_",  # 可选前缀
    #     digits=4         # 数字位数(如0001, 0002等)
    # )
    # exit()

    if deal_type==2: 
        # 处理已经分割好的文件
        file_path=pre_upload_path(deal_file_path)
    elif deal_type==1:
        if deal_file_name=="":
            print("deal_file_name文件路径不能为空")
            exit()
        if not os.path.exists(deal_file_name):
            print("deal_file_name文件不存在")
            exit()
        # file_path,upload_file_list=split_large_excel(deal_file_name)
        file_path,upload_file_list=split_excel_pd(deal_file_name)
        exit()
    
    else:
        print("deal_type参数错误")
        exit()
        
    page=login_to_website()
    if page==None:
        print("登录失败")
        exit()
    deal_files(page,file_path)
    # upload_file_to_website(page, r"D:\\develop\\huizixu\\temp_20250818_110004\\excel\\part_001.xlsx")